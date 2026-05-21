"""
11번가 아마존관 ASIN 수집 스크래퍼 (GitHub Actions용)
=========================================================
- 전체 카테고리 / 베스트 / 실시간구매 상품에서 ASIN + 상품명 수집
- 결과를 Excel (.xlsx) 로 저장 후 Google Drive 자동 업로드
- 환경변수: GDRIVE_CREDENTIALS (서비스 계정 JSON), GDRIVE_FOLDER_ID
"""

import requests
import re
import json
import os
import time
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ── 설정 ──────────────────────────────────────────────────
OUTPUT_FILE = "11st_amazon_asin.xlsx"

API_BASE     = "https://apis.11st.co.kr/pui/v2/page"
CATEGORY_API = "https://apis.11st.co.kr/display-api/apc/gnb/v1/categories"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":   "application/json",
    "Referer":  "https://www.11st.co.kr/amazon2",
    "Origin":   "https://www.11st.co.kr",
}

BEST_CTGR_NOS = [
    "0",
    "166153", "166154", "166156", "166157",
    "166158", "166159", "166160", "166162",
    "166163", "166164", "166165", "166166",
    "166182", "167628",
]

DELAY = 1.5   # 요청 간 딜레이(초)

# ── 데이터 수집 ────────────────────────────────────────────

def extract_asin(image_url: str) -> str:
    m = re.search(r"/asin/([A-Z0-9]{10})/", image_url or "")
    return m.group(1) if m else ""


def fetch_json(params: dict) -> dict | None:
    try:
        r = requests.get(API_BASE, params=params, headers=HEADERS, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [오류] params={params}: {e}")
        return None


def extract_products(data: dict, category: str = "") -> list[dict]:
    results = []

    def recurse(obj):
        if isinstance(obj, dict):
            img = obj.get("imageUrl", "")
            if img and "/asin/" in img:
                asin = extract_asin(img)
                if asin:
                    results.append({
                        "asin":          asin,
                        "product_name":  obj.get("prdNm") or obj.get("prdName") or "",
                        "prd_no":        str(obj.get("prdNo", "")),
                        "sell_price":    obj.get("sellPrice", ""),
                        "final_price":   obj.get("finalDscPrice", ""),
                        "discount_rate": obj.get("discountRate", ""),
                        "link_url":      obj.get("linkUrl", ""),
                        "category":      category,
                    })
                return
            for v in obj.values():
                recurse(v)
        elif isinstance(obj, list):
            for item in obj:
                recurse(item)

    recurse(data)
    return results


def collect_all() -> list[dict]:
    all_products = []

    # 1) 실시간구매
    print("[수집] 실시간구매")
    data = fetch_json({"pageId": "APCREALTIME"})
    if data:
        prods = extract_products(data, "실시간구매")
        print(f"  → {len(prods)}개")
        all_products.extend(prods)
    time.sleep(DELAY)

    # 2) 베스트 카테고리별
    for ctgr in BEST_CTGR_NOS:
        label = "전체" if ctgr == "0" else ctgr
        print(f"[수집] 베스트 ctgr1No={label}")
        data = fetch_json({"pageId": "APCBEST", "ctgr1No": ctgr})
        if data:
            prods = extract_products(data, f"베스트_{label}")
            print(f"  → {len(prods)}개")
            all_products.extend(prods)
        time.sleep(DELAY)

    # 3) 카테고리별
    try:
        r = requests.get(CATEGORY_API, headers=HEADERS, timeout=10)
        cats = r.json()
    except Exception as e:
        print(f"  [오류] 카테고리 목록: {e}")
        cats = []

    for cat in cats:
        cat_no   = cat.get("no")
        cat_name = cat.get("name", "")
        print(f"[수집] 카테고리: {cat_name}")
        data = fetch_json({"pageId": "APCCATEGORY", "dispCtgr1No": str(cat_no)})
        if data:
            prods = extract_products(data, cat_name)
            print(f"  → {len(prods)}개")
            all_products.extend(prods)
        time.sleep(DELAY)

    return all_products


# ── Excel 저장 ──────────────────────────────────────────────

def save_excel(products: list[dict], collected_at: str) -> int:
    # 중복 제거 (ASIN 기준)
    seen = {}
    for p in products:
        if p["asin"] not in seen:
            seen[p["asin"]] = p
    unique = list(seen.values())
    print(f"\n[저장] {len(products)}개 수집 → 중복 제거 후 {len(unique)}개")

    wb = Workbook()
    ws = wb.active
    ws.title = collected_at[:10]

    # 헤더
    orange = PatternFill("solid", fgColor="FF6C00")
    hfont  = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ("ASIN",         18), ("상품명",   55), ("카테고리", 20),
        ("판매가",       14), ("할인가",   14), ("할인율(%)", 12),
        ("11번가 링크",  55), ("상품번호", 18),
    ]
    for ci, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = hfont; cell.fill = orange
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 데이터
    for ri, p in enumerate(unique, 2):
        bg   = "FFFFFF" if ri % 2 == 0 else "FFF5EE"
        fill = PatternFill("solid", fgColor=bg)
        row  = [
            p["asin"], p["product_name"], p["category"],
            p["sell_price"], p["final_price"], p["discount_rate"],
            p["link_url"], p["prd_no"],
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [2, 7]))
            if ci == 7 and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    ws.cell(row=len(unique)+3, column=1, value=f"수집일시: {collected_at}")
    ws.cell(row=len(unique)+3, column=2, value=f"고유 ASIN: {len(unique)}개")

    wb.save(OUTPUT_FILE)
    print(f"[저장] {OUTPUT_FILE}")
    return len(unique)


# ── Google Drive 업로드 ────────────────────────────────────

def upload_to_drive(filename: str, collected_at: str) -> str:
    creds_json  = os.environ.get("GDRIVE_CREDENTIALS", "")
    folder_id   = os.environ.get("GDRIVE_FOLDER_ID", "")

    if not creds_json:
        print("[드라이브] GDRIVE_CREDENTIALS 환경변수 없음 — 업로드 건너뜀")
        return ""
    if not folder_id:
        print("[드라이브] GDRIVE_FOLDER_ID 환경변수 없음 — 업로드 건너뜀")
        return ""

    try:
        info = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/drive.file"]
        )
        service = build("drive", "v3", credentials=creds)

        date_str  = collected_at[:10].replace("-", "")
        drive_name = f"11st_amazon_asin_{date_str}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        file_meta = {"name": drive_name, "parents": [folder_id]}
        media     = MediaFileUpload(filename, mimetype=mime)
        result    = service.files().create(
            body=file_meta, media_body=media, fields="id,webViewLink"
        ).execute()

        link = result.get("webViewLink", "")
        print(f"[드라이브] 업로드 완료 → {drive_name}")
        print(f"  링크: {link}")
        return link

    except Exception as e:
        print(f"[드라이브] 업로드 오류: {e}")
        return ""


# ── 메인 ──────────────────────────────────────────────────

def main():
    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"=== 11번가 아마존관 ASIN 수집 시작: {collected_at} ===\n")

    products = collect_all()
    total    = save_excel(products, collected_at)
    link     = upload_to_drive(OUTPUT_FILE, collected_at)

    print(f"\n{'='*50}")
    print(f"수집 완료: 고유 ASIN {total}개")
    if link:
        print(f"드라이브: {link}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
